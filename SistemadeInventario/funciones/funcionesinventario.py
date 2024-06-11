from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os


#verifica si el producto esxiste en el diccionario o inventario

def verificarproducto(producto,diccionario):
    if producto.nameproducto in diccionario:
        print('es un producto')
        print()
        return 'existe'
    else:
       return 'validado'
    
    
#agregar un producto, si existe no lo agrega

def Listarproducto(producto,diccionario):
    verificarp = verificarproducto(producto,diccionario)
    if verificarp == 'existe':
        print('el producto no se puede agregar porque ya existe')
        print()
        return 'existe'
    else:
     diccionario[producto.nameproducto] = producto
     return 'valido'

def agregarUnidadesDeProducto(mensaje,productosdisponible):
       
       producto = input(mensaje)

       if producto in productosdisponible:
          
          producto = productosdisponible[producto]
          verificar = verificarproducto(producto,productosdisponible)

          if verificar == 'existe':
                cantidad = input('cantidad a agregar: ')
                producto.agregarunidades(int(cantidad))
                return ['valido',producto]
          else:
             return 'invalido'
          
       else:
           print('producto invalido o inexsistente')
           print()
           return 'invalido'
           
          
def restarUnidadesDeProducto(mensaje,productosdisponible):
       
       producto = input(mensaje)

       if producto in productosdisponible:
          
          producto = productosdisponible[producto]
          verificar = verificarproducto(producto,productosdisponible)

          if verificar == 'existe':
                cantidad = input('cantidad a restar: ')
                producto.restarunidades(int(cantidad))
                return ['valido',producto]
          else:
             return 'invalido'
       else:
           print('producto invalido o inexsistente')
           print()
           return 'invalido'
           

          
def mostrarUnidadesDeProducto(mensaje,productosdisponible):
       
       producto = input(mensaje)

       if producto in productosdisponible:
          
          producto = productosdisponible[producto]
          verificar = verificarproducto(producto,productosdisponible)

          if verificar == 'existe':
                return ['valido',producto]
          else:
             return 'invalido'
       else:
           print('producto invalido o inexsistente')
           print()
           return 'invalido'
           

           
          

def crearExcel(titulo, fecha, productosdisponible):
    # Obtener la ruta del escritorio del usuario
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Crear un nuevo libro de trabajo
    wb = Workbook()

    # Seleccionar la hoja activa
    ws = wb.active

    # Título y fecha
    titulo = titulo
    fecha_hoy = fecha

    # Escribir el título y la fecha
    ws.merge_cells('A1:D1')  # Merge cells para abarcar todas las columnas
    ws['A1'] = f'{titulo} - {fecha_hoy}'

    # Datos del diccionario
    productosdisponible = productosdisponible

    ws.append(['Producto', 'ID', 'Unidades', 'Precio'])
    for producto in productosdisponible.values():
        ws.append([producto.nameproducto, producto.iddeproducto, producto.unidades, producto.precio])

    # Ajustar el ancho de las columnas para que se ajuste automáticamente al contenido
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        column_letter = get_column_letter(cell.column)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar el libro de trabajo en el escritorio
    file_path = os.path.join(desktop_path, f"inventario_de_hoy_{fecha}.xlsx")
    wb.save(file_path)

    print(f"El archivo se ha guardado en: {file_path}")
    print()
    print('Por lo que se ha creado correctamente en el escritorio')
                
      
